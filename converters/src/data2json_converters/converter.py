# Copyright (c) 2026 Malte Doerper. MIT License. See LICENSE file.

from __future__ import annotations

import copy
import json
import re
import sys
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Callable, TextIO

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_SHEET_NAME = "data"
REVIEW_VERSION = "1"
REVIEW_ARTIFACT_MODE = "confirm-review"
LEGACY_REVIEW_ARTIFACT_MODE = "changes-review"
SOURCE_REF_COLUMNS = ("sourceType", "rawId", "key")
TIMESTAMP_PATTERN = re.compile(r"_(\d{4}-\d{2}-\d{2}-\d{2}-\d{2})$")
COMMENT_ENTRY_PATTERN = re.compile(r"^\[(?P<timestamp>[^\]]+)\]\s*<(?P<author>[^>]+)>:\s*(?P<text>.*)$")


def build_default_json_output_path(excel_path: Path, now: datetime | None = None) -> Path:
    timestamp = (now or datetime.now()).strftime("%Y-%m-%d-%H-%M")
    return excel_path.with_name(f"{excel_path.stem}_{timestamp}.json")


def build_default_excel_output_path(json_path: Path) -> Path:
    stem = TIMESTAMP_PATTERN.sub("", json_path.stem)
    return json_path.with_name(f"{stem}.xlsx")


def build_default_review_json_output_path(json_path: Path) -> Path:
    return json_path.with_name(f"{json_path.stem}-review.json")


def build_default_review_html_output_path(json_path: Path) -> Path:
    return json_path.with_name(f"{json_path.stem}-review.html")


def excel_to_json(
    excel_path: Path,
    *,
    json_path: Path | None = None,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> tuple[dict[str, list[dict[str, Any]]], list[str]]:
    warnings: list[str] = []
    items: list[dict[str, Any]] = []

    workbook = load_workbook(excel_path, data_only=True, read_only=True, keep_links=False)
    try:
        if sheet_name not in workbook.sheetnames:
            available_sheets = ", ".join(workbook.sheetnames) if workbook.sheetnames else "(none)"
            raise ValueError(
                f"Worksheet '{sheet_name}' not found in {excel_path}. "
                f"Available worksheets: {available_sheets}"
            )

        sheet = workbook[sheet_name]
        headers = _read_headers(sheet, warnings)

        for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            item = _excel_row_to_item(headers, row, row_index=row_index, warnings=warnings)
            if item:
                items.append(item)
    finally:
        workbook.close()

    payload = {"data": items}
    if json_path is not None:
        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    return payload, warnings


def json_to_excel(
    json_path: Path,
    *,
    excel_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    include_source_ref: bool = False,
) -> list[str]:
    payload = _load_dataset_payload(json_path)
    data = payload.get("data")
    if not isinstance(data, list):
        raise ValueError("JSON payload must contain a top-level 'data' array")

    warnings: list[str] = []
    _write_items_to_excel(
        data,
        excel_path=excel_path,
        sheet_name=sheet_name,
        include_source_ref=include_source_ref,
        warnings=warnings,
    )
    return warnings


def json_to_excel_changes(
    json_path: Path,
    *,
    excel_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    include_source_ref: bool = False,
) -> list[str]:
    payload = _load_dataset_payload(json_path)
    warnings: list[str] = []
    items = _overlay_changes_onto_data(payload, warnings=warnings)
    _write_items_to_excel(
        items,
        excel_path=excel_path,
        sheet_name=sheet_name,
        include_source_ref=include_source_ref,
        warnings=warnings,
    )
    return warnings


def json_to_excel_confirm_html(
    json_path: Path,
    *,
    excel_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    include_source_ref: bool = False,
    review_json_path: Path | None = None,
    review_html_path: Path | None = None,
) -> tuple[dict[str, str], list[str]]:
    review_artifact, warnings = build_review_artifact(
        json_path,
        excel_path=excel_path,
        sheet_name=sheet_name,
        include_source_ref=include_source_ref,
    )
    review_html_path = review_html_path or build_default_review_html_output_path(json_path)
    review_json_path = review_json_path or _build_review_json_path_for_html(review_html_path)
    result = save_review_artifact(review_artifact, review_json_path=review_json_path, review_html_path=review_html_path)
    return result, warnings


def _build_review_json_path_for_html(review_html_path: Path) -> Path:
    return review_html_path.with_suffix(".json")


def json_to_excel_confirm_cli(
    json_path: Path,
    *,
    excel_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    include_source_ref: bool = False,
    input_fn: Callable[[str], str] = input,
    output_stream: TextIO | None = None,
) -> list[str]:
    review_artifact, warnings = build_review_artifact(
        json_path,
        excel_path=excel_path,
        sheet_name=sheet_name,
        include_source_ref=include_source_ref,
    )
    _approve_review_rows_cli(review_artifact, input_fn=input_fn, output_stream=output_stream)
    warnings.extend(apply_review_artifact(review_artifact, excel_path=excel_path, sheet_name=sheet_name))
    return warnings


def apply_approved_review(
    review_json_path: Path,
    *,
    excel_path: Path | None = None,
    sheet_name: str | None = None,
) -> list[str]:
    review_artifact = json.loads(review_json_path.read_text(encoding="utf-8"))
    return apply_review_artifact(review_artifact, excel_path=excel_path, sheet_name=sheet_name)


def build_review_artifact(
    json_path: Path,
    *,
    excel_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    include_source_ref: bool = False,
) -> tuple[dict[str, Any], list[str]]:
    payload = _load_dataset_payload(json_path)
    warnings: list[str] = []
    review_artifact = _build_changes_review(
        payload,
        dataset_path=json_path,
        excel_path=excel_path,
        sheet_name=sheet_name,
        include_source_ref=include_source_ref,
        warnings=warnings,
    )
    return review_artifact, warnings


def save_review_artifact(
    review_artifact: dict[str, Any],
    *,
    review_json_path: Path,
    review_html_path: Path | None = None,
) -> dict[str, str]:
    review_artifact["reviewJsonPath"] = str(review_json_path)
    review_artifact["downloadFileName"] = review_json_path.name

    review_json_path.parent.mkdir(parents=True, exist_ok=True)
    review_json_path.write_text(json.dumps(review_artifact, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")

    result = {"review_json_path": str(review_json_path)}
    if review_html_path is not None:
        review_html_path.parent.mkdir(parents=True, exist_ok=True)
        review_html_path.write_text(_render_review_html(review_artifact), encoding="utf-8")
        result["review_html_path"] = str(review_html_path)
    return result


def apply_review_artifact(
    review_artifact: dict[str, Any],
    *,
    excel_path: Path | None = None,
    sheet_name: str | None = None,
) -> list[str]:
    return _apply_review_artifact(review_artifact, excel_path=excel_path, sheet_name=sheet_name)


def _write_items_to_excel(
    items: list[Any],
    *,
    excel_path: Path,
    sheet_name: str,
    include_source_ref: bool,
    warnings: list[str],
) -> None:
    workbook, sheet = _load_or_create_sheet(excel_path, sheet_name)
    headers = _ensure_headers(sheet, _derive_json_headers(items, include_source_ref=include_source_ref))
    header_index = {header: index + 1 for index, header in enumerate(headers)}
    row_lookup = _build_source_ref_lookup(sheet, header_index)
    item_id_lookup = _build_item_id_lookup(sheet, header_index)

    for item in items:
        if not isinstance(item, dict):
            warnings.append("Skipped non-object entry in data array")
            continue

        row_values = _item_to_excel_values(item, warnings, include_source_ref=include_source_ref)
        target_row = _find_target_row(item, row_lookup, item_id_lookup)
        if target_row is None:
            target_row = _next_data_row(sheet)
        _register_row_lookup(row_lookup, item_id_lookup, item, target_row)

        for header, value in row_values.items():
            column = header_index.get(header)
            if column is None:
                headers.append(header)
                column = len(headers)
                sheet.cell(row=1, column=column).value = header
                header_index[header] = column
            sheet.cell(row=target_row, column=column).value = value

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)


def _apply_review_artifact(
    review_artifact: dict[str, Any],
    *,
    excel_path: Path | None = None,
    sheet_name: str | None = None,
) -> list[str]:
    if review_artifact.get("mode") not in {REVIEW_ARTIFACT_MODE, LEGACY_REVIEW_ARTIFACT_MODE}:
        raise ValueError("Review artifact must have mode 'confirm-review' or legacy mode 'changes-review'")

    artifact_excel_path = review_artifact.get("excelPath")
    target_excel_path = excel_path or (Path(artifact_excel_path) if isinstance(artifact_excel_path, str) else None)
    if target_excel_path is None:
        raise ValueError("An Excel path is required to apply approved review rows")

    target_sheet_name = sheet_name or str(review_artifact.get("sheetName") or DEFAULT_SHEET_NAME)
    include_source_ref = bool(review_artifact.get("includeSourceRef"))

    warnings: list[str] = []
    workbook, sheet = _load_or_create_sheet(target_excel_path, target_sheet_name)
    headers = _ensure_headers(sheet, _read_existing_headers(sheet))
    header_index = {header: index + 1 for index, header in enumerate(headers)}

    rows = review_artifact.get("rows")
    if not isinstance(rows, list):
        raise ValueError("Review artifact must contain a top-level 'rows' array")

    for row in rows:
        if not isinstance(row, dict):
            warnings.append("Skipped malformed review row")
            continue

        change_id = str(row.get("changeId") or "<unknown>")
        if not row.get("approved"):
            continue

        status = str(row.get("status") or "invalid")
        if status != "ready":
            warnings.append(f"Skipped approved change {change_id} because status is {status}")
            continue

        action = str(row.get("action") or "")
        fields = row.get("fields")
        if not isinstance(fields, list):
            warnings.append(f"Skipped approved change {change_id} because it has no field payload")
            continue

        if action == "update":
            target = row.get("target")
            excel_row = target.get("excelRow") if isinstance(target, dict) else None
            if not isinstance(excel_row, int) or excel_row < 2:
                warnings.append(f"Skipped approved change {change_id} because it has no target Excel row")
                continue

            for field in fields:
                if not isinstance(field, dict) or not isinstance(field.get("name"), str):
                    warnings.append(f"Skipped malformed field in approved change {change_id}")
                    continue
                header = field["name"]
                column = header_index.get(header)
                if column is None:
                    headers.append(header)
                    column = len(headers)
                    sheet.cell(row=1, column=column).value = header
                    header_index[header] = column
                value = _serialize_cell_value(field.get("proposed"), key=header, warnings=warnings)
                sheet.cell(row=excel_row, column=column).value = value
            continue

        if action == "create":
            target = row.get("target") if isinstance(row.get("target"), dict) else {}
            item: dict[str, Any] = {}
            for field in fields:
                if not isinstance(field, dict) or not isinstance(field.get("name"), str):
                    warnings.append(f"Skipped malformed field in approved change {change_id}")
                    continue
                item[field["name"]] = field.get("proposed")
            if include_source_ref and isinstance(target.get("sourceRef"), dict):
                item["sourceRef"] = target["sourceRef"]

            row_values = _item_to_excel_values(item, warnings, include_source_ref=include_source_ref)
            target_row = _next_data_row(sheet)
            for header, value in row_values.items():
                column = header_index.get(header)
                if column is None:
                    headers.append(header)
                    column = len(headers)
                    sheet.cell(row=1, column=column).value = header
                    header_index[header] = column
                sheet.cell(row=target_row, column=column).value = value
            continue

        warnings.append(f"Skipped approved change {change_id} because action '{action}' is unsupported")

    target_excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target_excel_path)
    return warnings


def _overlay_changes_onto_data(payload: dict[str, Any], *, warnings: list[str]) -> list[dict[str, Any]]:
    data = payload.get("data")
    if not isinstance(data, list):
        raise ValueError("JSON payload must contain a top-level 'data' array")

    changes = payload.get("changes")
    if not isinstance(changes, dict):
        raise ValueError("Changes mode requires a top-level 'changes' object")

    change_rows = changes.get("rows")
    if not isinstance(change_rows, list):
        raise ValueError("Changes mode requires 'changes.rows' to be an array")

    items = [copy.deepcopy(item) for item in data if isinstance(item, dict)]
    skipped_entries = len(data) - len(items)
    if skipped_entries:
        warnings.extend(["Skipped non-object entry in data array"] * skipped_entries)

    source_lookup = _build_baseline_source_ref_lookup(items)
    item_id_lookup = _build_baseline_item_id_lookup(items)

    for change in change_rows:
        if not isinstance(change, dict):
            warnings.append("Skipped malformed change entry in changes.rows")
            continue

        change_id = str(change.get("changeId") or "<unknown>")
        action = str(change.get("action") or "")
        target = change.get("target") if isinstance(change.get("target"), dict) else {}
        baseline = change.get("baseline") if isinstance(change.get("baseline"), dict) else None
        proposed = change.get("proposed") if isinstance(change.get("proposed"), dict) else None
        source_ref = target.get("sourceRef") if isinstance(target.get("sourceRef"), dict) else None
        item_id = _normalize_item_id(target.get("itemId"))

        if baseline is None or proposed is None:
            warnings.append(f"Change {change_id} must provide object values for baseline and proposed")
            continue

        if action == "update":
            item = _find_baseline_item(target, source_lookup, item_id_lookup)
            if item is None:
                warnings.append(f"Change {change_id} could not find a matching baseline item")
                continue

            for field_name, declared_baseline in baseline.items():
                actual_baseline = item.get(field_name)
                if declared_baseline != actual_baseline:
                    warnings.append(
                        f"Change {change_id} baseline for field '{field_name}' does not match baseline data"
                    )

            for field_name, proposed_value in proposed.items():
                item[field_name] = copy.deepcopy(proposed_value)

            if item_id is not None and _normalize_item_id(item.get("id")) is None:
                item["id"] = item_id
            if source_ref and not isinstance(item.get("sourceRef"), dict):
                item["sourceRef"] = copy.deepcopy(source_ref)
            continue

        if action == "create":
            if baseline:
                warnings.append(f"Change {change_id} create rows should use an empty baseline object")

            existing_item = _find_baseline_item(target, source_lookup, item_id_lookup)
            if existing_item is not None:
                warnings.append(f"Change {change_id} create target already exists in baseline data")
                continue

            item = copy.deepcopy(proposed)
            if item_id is not None and _normalize_item_id(item.get("id")) is None:
                item["id"] = item_id
            if source_ref:
                item["sourceRef"] = copy.deepcopy(source_ref)

            items.append(item)
            lookup_key = _source_ref_lookup_key(item.get("sourceRef") if isinstance(item.get("sourceRef"), dict) else None)
            if lookup_key is not None:
                source_lookup[lookup_key] = item
            created_item_id = _normalize_item_id(item.get("id"))
            if created_item_id is not None:
                item_id_lookup[created_item_id] = item
            continue

        warnings.append(f"Change {change_id} uses unsupported action '{action}'")

    return items


def _approve_review_rows_cli(
    review_artifact: dict[str, Any],
    *,
    input_fn: Callable[[str], str],
    output_stream: TextIO | None,
) -> None:
    rows = review_artifact.get("rows")
    if not isinstance(rows, list):
        raise ValueError("Review artifact must contain a top-level 'rows' array")

    stream = output_stream or sys.stdout
    print("Review pending Excel changes", file=stream)
    print(f"Dataset: {review_artifact.get('datasetPath', '')}", file=stream)
    print(f"Workbook: {review_artifact.get('excelPath', '')}", file=stream)
    print("", file=stream)

    approved_count = 0
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue

        change_id = str(row.get("changeId") or f"change-{index}")
        action = str(row.get("action") or "unknown")
        status = str(row.get("status") or "invalid")
        target = row.get("target") if isinstance(row.get("target"), dict) else {}

        print(f"[{index}/{len(rows)}] {change_id} ({action})", file=stream)
        print(f"  status: {status}", file=stream)
        print(f"  itemId: {target.get('itemId', '')}", file=stream)
        print(f"  excelRow: {target.get('excelRow', '')}", file=stream)
        if target.get("sourceRef"):
            print(f"  sourceRef: {json.dumps(target['sourceRef'], ensure_ascii=True)}", file=stream)

        fields = row.get("fields") if isinstance(row.get("fields"), list) else []
        for field in fields:
            if not isinstance(field, dict):
                continue
            print(f"    field: {field.get('name', '')}", file=stream)
            print(f"      excel: {_format_review_value(field.get('excelCurrent'))}", file=stream)
            print(f"      baseline: {_format_review_value(field.get('jsonBaseline'))}", file=stream)
            print(f"      proposed: {_format_review_value(field.get('proposed'))}", file=stream)
            print(f"      conflict: {'yes' if field.get('conflict') else 'no'}", file=stream)

        if status != "ready":
            row["approved"] = False
            print(f"  skipping prompt because status is {status}", file=stream)
            print("", file=stream)
            continue

        answer = input_fn(f"Approve {change_id}? [y/N]: ").strip().lower()
        row["approved"] = answer in {"y", "yes"}
        if row["approved"]:
            approved_count += 1
            print("  approved", file=stream)
        else:
            print("  skipped", file=stream)
        print("", file=stream)

    print(f"Approved {approved_count} of {len(rows)} change(s).", file=stream)


def _format_review_value(value: Any) -> str:
    if value is None:
        return "<empty>"
    if value == "":
        return "<empty>"
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=True)


def _build_changes_review(
    payload: dict[str, Any],
    *,
    dataset_path: Path,
    excel_path: Path,
    sheet_name: str,
    include_source_ref: bool,
    warnings: list[str],
) -> dict[str, Any]:
    data = payload.get("data")
    if not isinstance(data, list):
        raise ValueError("JSON payload must contain a top-level 'data' array")

    changes = payload.get("changes")
    if not isinstance(changes, dict):
        raise ValueError("Changes mode requires a top-level 'changes' object")

    change_rows = changes.get("rows")
    if not isinstance(change_rows, list):
        raise ValueError("Changes mode requires 'changes.rows' to be an array")

    workbook, sheet = _load_or_create_sheet(excel_path, sheet_name)
    existing_headers = _read_existing_headers(sheet)
    header_index = {header: index + 1 for index, header in enumerate(existing_headers)}
    source_ref_lookup = _build_source_ref_lookup(sheet, header_index)
    item_id_lookup = _build_item_id_lookup(sheet, header_index)
    baseline_source_lookup = _build_baseline_source_ref_lookup(data)
    baseline_item_id_lookup = _build_baseline_item_id_lookup(data)

    review_rows = []
    for change in change_rows:
        review_rows.append(
            _build_review_row(
                change,
                sheet=sheet,
                header_index=header_index,
                source_ref_lookup=source_ref_lookup,
                item_id_lookup=item_id_lookup,
                baseline_source_lookup=baseline_source_lookup,
                baseline_item_id_lookup=baseline_item_id_lookup,
                warnings=warnings,
            )
        )

    return {
        "version": REVIEW_VERSION,
        "mode": REVIEW_ARTIFACT_MODE,
        "datasetPath": str(dataset_path),
        "excelPath": str(excel_path),
        "sheetName": sheet_name,
        "includeSourceRef": include_source_ref,
        "generatedAt": _isoformat_datetime(datetime.now(UTC)),
        "warnings": list(warnings),
        "rows": review_rows,
    }


def _build_review_row(
    change: Any,
    *,
    sheet: Worksheet,
    header_index: dict[str, int],
    source_ref_lookup: dict[tuple[tuple[str, Any], ...], int],
    item_id_lookup: dict[str, int],
    baseline_source_lookup: dict[tuple[tuple[str, Any], ...], dict[str, Any]],
    baseline_item_id_lookup: dict[str, dict[str, Any]],
    warnings: list[str],
) -> dict[str, Any]:
    if not isinstance(change, dict):
        warnings.append("Skipped malformed change entry in changes.rows")
        return {
            "changeId": "",
            "action": "",
            "approved": False,
            "status": "invalid",
            "target": {},
            "fields": [],
        }

    change_id = str(change.get("changeId") or "")
    action = str(change.get("action") or "")
    target = change.get("target") if isinstance(change.get("target"), dict) else {}
    baseline = change.get("baseline") if isinstance(change.get("baseline"), dict) else None
    proposed = change.get("proposed") if isinstance(change.get("proposed"), dict) else None
    review_target = _copy_target(target)

    review_row = {
        "changeId": change_id,
        "action": action,
        "approved": False,
        "status": "ready",
        "target": review_target,
        "fields": [],
    }

    if action not in {"update", "create"}:
        warnings.append(f"Change {change_id or '<unknown>'} uses unsupported action '{action}'")
        review_row["status"] = "invalid"
        return review_row

    if baseline is None or proposed is None:
        warnings.append(f"Change {change_id or '<unknown>'} must provide object values for baseline and proposed")
        review_row["status"] = "invalid"
        return review_row

    source_ref = target.get("sourceRef") if isinstance(target.get("sourceRef"), dict) else None
    item_id = _normalize_item_id(target.get("itemId"))

    if action == "update":
        if not source_ref and item_id is None:
            warnings.append(f"Change {change_id or '<unknown>'} is missing both target.sourceRef and target.itemId")
            review_row["status"] = "invalid"
            return review_row

        baseline_item = _find_baseline_item(target, baseline_source_lookup, baseline_item_id_lookup)
        if baseline_item is None:
            warnings.append(f"Change {change_id or '<unknown>'} could not find a matching baseline item")
            review_row["status"] = "invalid"
            return review_row

        field_names = _ordered_field_names(baseline, proposed)

        if not field_names:
            warnings.append(f"Change {change_id or '<unknown>'} has no changed fields")
            review_row["status"] = "invalid"

        excel_row = _find_target_row_for_change(target, source_ref_lookup, item_id_lookup)
        if excel_row is None and review_row["status"] == "ready":
            review_row["status"] = "unresolved"
        if excel_row is not None:
            review_target["excelRow"] = excel_row

        any_conflict = False
        for field_name in field_names:
            actual_baseline = baseline_item.get(field_name)
            declared_baseline = baseline.get(field_name)
            if declared_baseline != actual_baseline:
                warnings.append(
                    f"Change {change_id or '<unknown>'} baseline for field '{field_name}' does not match baseline data"
                )
                review_row["status"] = "invalid"

            excel_current = _read_sheet_value(
                sheet,
                header_index,
                excel_row,
                field_name,
                warnings=warnings,
            )
            conflict = excel_row is not None and excel_current != actual_baseline
            any_conflict = any_conflict or conflict
            review_row["fields"].append(
                {
                    "name": field_name,
                    "excelCurrent": excel_current,
                    "jsonBaseline": actual_baseline,
                    "proposed": proposed.get(field_name),
                    "conflict": conflict,
                }
            )

        if review_row["status"] == "ready" and any_conflict:
            review_row["status"] = "conflict"
        return review_row

    if baseline:
        warnings.append(f"Change {change_id or '<unknown>'} create rows must use an empty baseline object")
        review_row["status"] = "invalid"
    if item_id is None:
        warnings.append(f"Change {change_id or '<unknown>'} create rows must include target.itemId")
        review_row["status"] = "invalid"

    field_names = list(proposed.keys())
    if not field_names:
        warnings.append(f"Change {change_id or '<unknown>'} create row has no proposed fields")
        review_row["status"] = "invalid"

    existing_row = None
    if source_ref:
        existing_row = source_ref_lookup.get(_source_ref_lookup_key(source_ref))
        if existing_row is not None and review_row["status"] == "ready":
            review_row["status"] = "conflict"
            review_target["excelRow"] = existing_row

    for field_name in field_names:
        excel_current = _read_sheet_value(sheet, header_index, existing_row, field_name, warnings=warnings)
        review_row["fields"].append(
            {
                "name": field_name,
                "excelCurrent": excel_current,
                "jsonBaseline": None,
                "proposed": proposed.get(field_name),
                "conflict": existing_row is not None,
            }
        )

    return review_row


def _render_review_html(review_artifact: dict[str, Any]) -> str:
    review_json = json.dumps(review_artifact, ensure_ascii=True).replace("</", "<\\/")
    title = "Changes Review"
    return f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
  <meta charset=\"utf-8\">
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">
  <title>{title}</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f7f4ed;
      --panel: #fffdf8;
      --border: #d8cdb9;
      --text: #2f2419;
      --muted: #6c5a48;
      --accent: #1d6f5f;
      --conflict: #b5462a;
      --invalid: #8b2e3e;
      --shadow: rgba(47, 36, 25, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Georgia, "Times New Roman", serif;
      background:
        radial-gradient(circle at top left, rgba(29, 111, 95, 0.14), transparent 32%),
        linear-gradient(180deg, #fbf8f2 0%, var(--bg) 100%);
      color: var(--text);
    }}
    main {{ max-width: 1180px; margin: 0 auto; padding: 32px 20px 64px; }}
    .hero {{ display: grid; gap: 12px; margin-bottom: 24px; }}
    .hero h1 {{ margin: 0; font-size: clamp(2rem, 4vw, 3rem); line-height: 1; }}
    .hero p {{ margin: 0; color: var(--muted); max-width: 75ch; }}
    .toolbar {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
      align-items: center;
      justify-content: space-between;
      padding: 16px 18px;
      margin-bottom: 18px;
      border: 1px solid var(--border);
      border-radius: 18px;
      background: rgba(255, 253, 248, 0.95);
      box-shadow: 0 12px 32px var(--shadow);
    }}
    .toolbar button {{
      border: 0;
      border-radius: 999px;
      padding: 10px 18px;
      background: var(--accent);
      color: white;
      font: inherit;
      cursor: pointer;
    }}
    .toolbar .meta {{ color: var(--muted); font-size: 0.95rem; }}
    .row-card {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-left-width: 8px;
      border-radius: 18px;
      box-shadow: 0 12px 30px var(--shadow);
      margin-bottom: 18px;
      overflow: hidden;
    }}
    .row-card.ready {{ border-left-color: var(--accent); }}
    .row-card.conflict {{ border-left-color: var(--conflict); }}
    .row-card.unresolved, .row-card.invalid {{ border-left-color: var(--invalid); }}
    .row-header {{
      display: flex;
      flex-wrap: wrap;
      gap: 14px;
      align-items: center;
      justify-content: space-between;
      padding: 18px 20px;
      border-bottom: 1px solid var(--border);
      background: rgba(216, 205, 185, 0.18);
    }}
    .row-title {{ display: grid; gap: 6px; }}
    .row-title strong {{ font-size: 1.15rem; }}
    .badges {{ display: flex; gap: 8px; flex-wrap: wrap; }}
    .badge {{
      display: inline-flex;
      align-items: center;
      border-radius: 999px;
      padding: 4px 10px;
      font-size: 0.85rem;
      background: rgba(29, 111, 95, 0.1);
      color: var(--accent);
    }}
    .badge.conflict {{ background: rgba(181, 70, 42, 0.12); color: var(--conflict); }}
    .badge.invalid {{ background: rgba(139, 46, 62, 0.12); color: var(--invalid); }}
    .approve {{ display: inline-flex; gap: 8px; align-items: center; font-weight: 600; }}
    .body {{ padding: 18px 20px 22px; display: grid; gap: 16px; }}
    .identity {{ color: var(--muted); font-size: 0.95rem; display: grid; gap: 4px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ text-align: left; vertical-align: top; padding: 10px 12px; border-bottom: 1px solid var(--border); }}
    th {{ font-size: 0.82rem; letter-spacing: 0.04em; text-transform: uppercase; color: var(--muted); }}
    td code {{ font-family: Consolas, monospace; white-space: pre-wrap; word-break: break-word; }}
    .field-conflict {{ background: rgba(181, 70, 42, 0.08); }}
    .empty {{ color: #97836f; font-style: italic; }}
    @media (max-width: 720px) {{
      .row-header {{ align-items: flex-start; }}
      th:nth-child(1), td:nth-child(1) {{ position: sticky; left: 0; background: var(--panel); }}
    }}
  </style>
</head>
<body>
  <main>
    <section class=\"hero\">
      <h1>Review Pending Excel Changes</h1>
            <p>Approve rows in the browser, then download the updated review JSON and re-run <code>json-to-excel --mode confirm-html --apply-approved &lt;review.json&gt;</code> to write approved rows only.</p>
    </section>
    <section class=\"toolbar\">
      <div class=\"meta\" id=\"summary\"></div>
            <button type=\"button\" id=\"save-json\">Save approved review JSON</button>
    </section>
    <section id=\"rows\"></section>
  </main>
  <script id=\"review-data\" type=\"application/json\">{review_json}</script>
  <script>
        const reviewArtifact = JSON.parse(document.getElementById("review-data").textContent);
    const rowsContainer = document.getElementById("rows");
    const summary = document.getElementById("summary");
        const saveJsonButton = document.getElementById("save-json");
        const supportsSavePicker = typeof window.showSaveFilePicker === "function";
        const supportsHandlePersistence = supportsSavePicker && typeof indexedDB !== "undefined";
        const handleDbName = "data2json-converters-review-handles";
        const handleStoreName = "reviewFiles";
        const handleKey = `review-json::${{reviewArtifact.reviewJsonPath || reviewArtifact.downloadFileName || "confirm-review.json"}}`;
        let handleDbPromise = null;
        let cachedReviewHandle = null;

    function formatValue(value) {{
      if (value === null || value === undefined || value === "") {{
        return "";
      }}
      return typeof value === "string" ? value : JSON.stringify(value, null, 2);
    }}

        function buildSummaryText(extraMessage = "") {{
            const approvedCount = reviewArtifact.rows.filter((row) => row.approved).length;
            const saveMode = supportsSavePicker
                ? "Save button remembers the chosen review JSON file in this browser."
                : "Browser save-handle support is unavailable; fallback uses a download.";
            const base = `${{reviewArtifact.rows.length}} pending row(s), ${{approvedCount}} approved. Dataset: ${{reviewArtifact.datasetPath}}. Workbook: ${{reviewArtifact.excelPath}}. ${{saveMode}}`;
            return extraMessage ? `${{base}} ${{extraMessage}}` : base;
        }}

    function updateSummary() {{
            summary.textContent = buildSummaryText();
    }}

        async function openHandleDb() {{
            if (!supportsHandlePersistence) {{
                return null;
            }}
            if (!handleDbPromise) {{
                handleDbPromise = new Promise((resolve, reject) => {{
                    const request = indexedDB.open(handleDbName, 1);
                    request.onupgradeneeded = () => {{
                        const db = request.result;
                        if (!db.objectStoreNames.contains(handleStoreName)) {{
                            db.createObjectStore(handleStoreName);
                        }}
                    }};
                    request.onsuccess = () => resolve(request.result);
                    request.onerror = () => reject(request.error || new Error("Failed to open review handle database"));
                }});
            }}
            return handleDbPromise;
        }}

        async function readPersistedHandle() {{
            if (!supportsHandlePersistence) {{
                return null;
            }}
            const db = await openHandleDb();
            return new Promise((resolve, reject) => {{
                const transaction = db.transaction(handleStoreName, "readonly");
                const store = transaction.objectStore(handleStoreName);
                const request = store.get(handleKey);
                request.onsuccess = () => resolve(request.result || null);
                request.onerror = () => reject(request.error || new Error("Failed to read persisted review handle"));
            }});
        }}

        async function persistHandle(handle) {{
            if (!supportsHandlePersistence) {{
                return;
            }}
            const db = await openHandleDb();
            await new Promise((resolve, reject) => {{
                const transaction = db.transaction(handleStoreName, "readwrite");
                const store = transaction.objectStore(handleStoreName);
                const request = store.put(handle, handleKey);
                request.onsuccess = () => resolve();
                request.onerror = () => reject(request.error || new Error("Failed to persist review handle"));
            }});
        }}

        async function getCachedReviewHandle() {{
            if (cachedReviewHandle) {{
                return cachedReviewHandle;
            }}
            cachedReviewHandle = await readPersistedHandle();
            return cachedReviewHandle;
        }}

        async function setCachedReviewHandle(handle) {{
            cachedReviewHandle = handle;
            await persistHandle(handle);
        }}

        async function ensureWritePermission(handle) {{
            if (!handle || typeof handle.requestPermission !== "function") {{
                return true;
            }}
            if (typeof handle.queryPermission === "function") {{
                const permission = await handle.queryPermission({{ mode: "readwrite" }});
                if (permission === "granted") {{
                    return true;
                }}
            }}
            const permission = await handle.requestPermission({{ mode: "readwrite" }});
            return permission === "granted";
        }}

        async function writeReviewToHandle(handle) {{
            const writable = await handle.createWritable();
            await writable.write(JSON.stringify(reviewArtifact, null, 2) + "\\n");
            await writable.close();
            return handle.name || reviewArtifact.downloadFileName || "confirm-review.json";
        }}

        async function saveReviewWithPicker() {{
            const existingHandle = await getCachedReviewHandle();
            if (existingHandle && await ensureWritePermission(existingHandle)) {{
                return writeReviewToHandle(existingHandle);
            }}

            const fileHandle = await window.showSaveFilePicker({{
                id: "data2json-converters-review-json",
                suggestedName: reviewArtifact.downloadFileName || "confirm-review.json",
                types: [{{
                    description: "JSON files",
                    accept: {{ "application/json": [".json"] }},
                }}],
            }});
            await setCachedReviewHandle(fileHandle);
            return writeReviewToHandle(fileHandle);
        }}

        function saveReviewWithDownload() {{
            const blob = new Blob([JSON.stringify(reviewArtifact, null, 2) + "\\n"], {{ type: "application/json" }});
            const url = URL.createObjectURL(blob);
            const anchor = document.createElement("a");
            anchor.href = url;
            anchor.download = reviewArtifact.downloadFileName || "confirm-review.json";
            anchor.click();
            setTimeout(() => URL.revokeObjectURL(url), 500);
            return anchor.download;
        }}

        async function saveApprovedReviewJson() {{
            saveJsonButton.disabled = true;
            try {{
                const savedName = supportsSavePicker
                    ? await saveReviewWithPicker()
                    : saveReviewWithDownload();
                if (savedName) {{
                    summary.textContent = buildSummaryText(`Saved approvals to ${{savedName}}.`);
                }} else {{
                    updateSummary();
                }}
            }} catch (error) {{
                console.error("Failed to save approved review JSON", error);
                summary.textContent = buildSummaryText(`Save failed: ${{error.message || error}}`);
            }} finally {{
                saveJsonButton.disabled = false;
            }}
        }}

    function renderRow(row) {{
      const article = document.createElement("article");
      article.className = `row-card ${{row.status || "invalid"}}`;

      const header = document.createElement("div");
      header.className = "row-header";

      const title = document.createElement("div");
      title.className = "row-title";
      const strong = document.createElement("strong");
      strong.textContent = row.changeId || "Unnamed change";
      title.appendChild(strong);

      const badges = document.createElement("div");
      badges.className = "badges";
      const actionBadge = document.createElement("span");
      actionBadge.className = "badge";
      actionBadge.textContent = row.action || "unknown";
      badges.appendChild(actionBadge);

      const statusBadge = document.createElement("span");
      statusBadge.className = `badge ${{row.status === "ready" ? "" : row.status}}`.trim();
      statusBadge.textContent = row.status || "invalid";
      badges.appendChild(statusBadge);
      title.appendChild(badges);

      const approve = document.createElement("label");
      approve.className = "approve";
      const checkbox = document.createElement("input");
      checkbox.type = "checkbox";
      checkbox.checked = Boolean(row.approved);
      checkbox.disabled = row.status !== "ready";
      checkbox.addEventListener("change", () => {{
        row.approved = checkbox.checked;
        updateSummary();
      }});
      approve.appendChild(checkbox);
      approve.appendChild(document.createTextNode("Approve row"));

      header.appendChild(title);
      header.appendChild(approve);
      article.appendChild(header);

      const body = document.createElement("div");
      body.className = "body";
      const identity = document.createElement("div");
      identity.className = "identity";
      const target = row.target || {{}};
      identity.innerHTML = `
        <div><strong>Item:</strong> ${{target.itemId || ""}}</div>
        <div><strong>SourceRef:</strong> ${{target.sourceRef ? JSON.stringify(target.sourceRef) : ""}}</div>
        <div><strong>Excel row:</strong> ${{target.excelRow || ""}}</div>
      `;
      body.appendChild(identity);

      const table = document.createElement("table");
      table.innerHTML = `
        <thead>
          <tr>
            <th>Field</th>
            <th>Current Excel value</th>
            <th>Unchanged JSON baseline</th>
            <th>Proposed changed value</th>
            <th>Conflict</th>
          </tr>
        </thead>
      `;
      const tbody = document.createElement("tbody");
      for (const field of row.fields || []) {{
        const tr = document.createElement("tr");
        if (field.conflict) {{
          tr.className = "field-conflict";
        }}
        const values = [
          field.name,
          formatValue(field.excelCurrent),
          formatValue(field.jsonBaseline),
          formatValue(field.proposed),
          field.conflict ? "yes" : "no",
        ];
        for (const value of values) {{
          const td = document.createElement("td");
          const code = document.createElement("code");
          if (value === "") {{
            code.className = "empty";
            code.textContent = "empty";
          }} else {{
            code.textContent = value;
          }}
          td.appendChild(code);
          tr.appendChild(td);
        }}
        tbody.appendChild(tr);
      }}
      table.appendChild(tbody);
      body.appendChild(table);
      article.appendChild(body);
      return article;
    }}

    function render() {{
      rowsContainer.innerHTML = "";
      for (const row of reviewArtifact.rows || []) {{
        rowsContainer.appendChild(renderRow(row));
      }}
      updateSummary();
    }}

        saveJsonButton.addEventListener("click", () => {{
            void saveApprovedReviewJson();
        }});

    render();
  </script>
</body>
</html>
"""


def _load_dataset_payload(json_path: Path) -> dict[str, Any]:
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("JSON payload must be a top-level object")
    return payload


def _load_or_create_sheet(excel_path: Path, sheet_name: str) -> tuple[Workbook, Worksheet]:
    if excel_path.exists():
        workbook = load_workbook(excel_path)
    else:
        workbook = Workbook()

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        if workbook.sheetnames == ["Sheet"] and workbook["Sheet"].max_row == 1 and workbook["Sheet"].max_column == 1:
            sheet = workbook["Sheet"]
            sheet.title = sheet_name
        else:
            sheet = workbook.create_sheet(title=sheet_name)

    return workbook, sheet


def _read_headers(sheet: Worksheet, warnings: list[str]) -> list[str | None]:
    headers: list[str | None] = []
    for column_index, cell in enumerate(_iter_header_cells(sheet), start=1):
        raw_value = cell.value
        if raw_value is None or str(raw_value).strip() == "":
            warnings.append(f"Ignoring blank header in column {column_index}")
            headers.append(None)
            continue
        headers.append(str(raw_value).strip())
    return headers


def _read_existing_headers(sheet: Worksheet) -> list[str]:
    headers: list[str] = []
    for cell in _iter_header_cells(sheet):
        if cell.value in (None, ""):
            continue
        headers.append(str(cell.value).strip())
    return headers


def _iter_header_cells(sheet: Worksheet) -> tuple[Any, ...]:
    header_rows = sheet.iter_rows(min_row=1, max_row=1)
    return next(header_rows, ())


def _excel_row_to_item(
    headers: list[str | None],
    row: tuple[Any, ...],
    *,
    row_index: int,
    warnings: list[str],
) -> dict[str, Any]:
    item: dict[str, Any] = {}
    source_ref: dict[str, Any] = {}

    for header, cell in zip(headers, row):
        if header is None:
            continue

        cell_value = cell.value if hasattr(cell, "value") else cell
        number_format = cell.number_format if hasattr(cell, "number_format") else None
        normalized = _normalize_excel_value(
            cell_value,
            field_name=header,
            row_index=row_index,
            warnings=warnings,
            number_format=number_format,
        )
        if normalized is None:
            continue

        normalized_header = "comment" if header == "comments" else header

        if header in SOURCE_REF_COLUMNS:
            source_ref[header] = normalized
        else:
            item[normalized_header] = normalized

    if source_ref:
        item["sourceRef"] = source_ref
    return item


def _normalize_excel_value(
    value: Any,
    *,
    field_name: str,
    row_index: int,
    warnings: list[str],
    number_format: str | None = None,
) -> Any:
    if value is None:
        return None

    if isinstance(value, datetime):
        return _normalize_excel_datetime_value(value, number_format=number_format)

    if isinstance(value, date):
        return _isoformat_date_value(value)

    if isinstance(value, str):
        trimmed = value.strip()
        if trimmed == "":
            return None

        comment_value = _parse_comment_cell(trimmed, field_name=field_name, row_index=row_index, warnings=warnings)
        if comment_value is not None:
            return comment_value

        multi_value = _parse_multi_value_cell(trimmed)
        if multi_value is not None:
            return multi_value

        normalized_timestamp = _normalize_timestamp_string(trimmed)
        return normalized_timestamp if normalized_timestamp is not None else trimmed

    return value


def _parse_multi_value_cell(value: str) -> list[str] | None:
    if ";" not in value:
        return None
    parts = [part.strip() for part in value.split(";") if part.strip()]
    if len(parts) <= 1:
        return None
    return parts


def _parse_comment_cell(value: str, *, field_name: str, row_index: int, warnings: list[str]) -> dict[str, Any] | None:
    entries = [entry.strip() for entry in value.split(";") if entry.strip()]
    if not entries:
        return None

    has_comment_shape = any(entry.startswith("[") and "<" in entry and ">:" in entry for entry in entries)
    if not has_comment_shape:
        return None

    messages: list[dict[str, str]] = []
    for entry in entries:
        match = COMMENT_ENTRY_PATTERN.match(entry)
        if not match:
            warnings.append(f"Row {row_index} field '{field_name}' contains malformed comment entry")
            messages.append({"author": "", "timestamp": "", "text": entry})
            continue

        raw_timestamp = match.group("timestamp").strip()
        normalized_timestamp = _normalize_timestamp_string(raw_timestamp)
        if normalized_timestamp is None:
            warnings.append(
                f"Row {row_index} field '{field_name}' contains an unparseable comment timestamp: {raw_timestamp}"
            )
            normalized_timestamp = raw_timestamp

        messages.append(
            {
                "author": match.group("author").strip(),
                "timestamp": normalized_timestamp,
                "text": match.group("text").strip(),
            }
        )

    return {"threads": [{"id": "t1", "messages": messages}]}


def _normalize_timestamp_string(value: str) -> str | None:
    for parsed in _iter_datetime_candidates(value):
        return _isoformat_datetime(parsed)
    return None


def _normalize_excel_datetime_value(value: datetime, *, number_format: str | None = None) -> str:
    if _is_excel_date_only_format(number_format):
        return _isoformat_date_value(value.date())
    return _isoformat_datetime(value)


def _isoformat_date_value(value: date) -> str:
    return value.isoformat()


def _is_excel_date_only_format(number_format: str | None) -> bool:
    if not number_format:
        return False

    normalized = number_format.lower()

    # Strip quoted literals, locale/color blocks, and escaped characters.
    normalized = re.sub(r'"[^"]*"', '', normalized)
    normalized = re.sub(r'\[[^\]]*\]', '', normalized)
    normalized = re.sub(r'\\.', '', normalized)

    return 'h' not in normalized and 's' not in normalized and 'am/pm' not in normalized


def _iter_datetime_candidates(value: str):
    text = value.strip()
    if not text:
        return

    iso_candidate = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(iso_candidate)
    except ValueError:
        parsed = None
    if parsed is not None:
        yield parsed
        return

    formats = (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%y",
        "%m/%d/%y %H:%M",
        "%m/%d/%y %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
    )
    for fmt in formats:
        try:
            yield datetime.strptime(text, fmt)
            return
        except ValueError:
            continue


def _isoformat_datetime(value: datetime) -> str:
    return value.date().isoformat()


def _derive_json_headers(items: list[Any], *, include_source_ref: bool) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        for key, value in item.items():
            if key == "sourceRef":
                if include_source_ref and isinstance(value, dict):
                    for source_key in SOURCE_REF_COLUMNS:
                        if source_key in value and source_key not in seen:
                            headers.append(source_key)
                            seen.add(source_key)
                continue
            normalized_key = "comment" if key == "comments" else key
            if normalized_key not in seen:
                headers.append(normalized_key)
                seen.add(normalized_key)
    return headers


def _ensure_headers(sheet: Worksheet, desired_headers: list[str]) -> list[str]:
    headers = list(_read_existing_headers(sheet))
    for header in desired_headers:
        if header not in headers:
            headers.append(header)

    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column_index).value = header
    return headers


def _build_source_ref_lookup(sheet: Worksheet, header_index: dict[str, int]) -> dict[tuple[tuple[str, Any], ...], int]:
    lookup: dict[tuple[tuple[str, Any], ...], int] = {}
    present_source_columns = [key for key in SOURCE_REF_COLUMNS if key in header_index]
    if not present_source_columns:
        return lookup

    for row_index in range(2, sheet.max_row + 1):
        source_ref: dict[str, Any] = {}
        for key in present_source_columns:
            value = sheet.cell(row=row_index, column=header_index[key]).value
            normalized = _normalize_lookup_value(value)
            if normalized is not None:
                source_ref[key] = normalized
        lookup_key = _source_ref_lookup_key(source_ref)
        if lookup_key is not None:
            lookup[lookup_key] = row_index
    return lookup


def _build_item_id_lookup(sheet: Worksheet, header_index: dict[str, int]) -> dict[str, int]:
    if "id" not in header_index:
        return {}

    lookup: dict[str, int] = {}
    id_column = header_index["id"]
    for row_index in range(2, sheet.max_row + 1):
        item_id = _normalize_item_id(sheet.cell(row=row_index, column=id_column).value)
        if item_id is not None:
            lookup[item_id] = row_index
    return lookup


def _build_baseline_source_ref_lookup(items: list[Any]) -> dict[tuple[tuple[str, Any], ...], dict[str, Any]]:
    lookup: dict[tuple[tuple[str, Any], ...], dict[str, Any]] = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        source_ref = item.get("sourceRef")
        lookup_key = _source_ref_lookup_key(source_ref if isinstance(source_ref, dict) else None)
        if lookup_key is not None:
            lookup[lookup_key] = item
    return lookup


def _build_baseline_item_id_lookup(items: list[Any]) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        item_id = _normalize_item_id(item.get("id"))
        if item_id is not None:
            lookup[item_id] = item
    return lookup


def _find_target_row(
    item: dict[str, Any],
    source_ref_lookup: dict[tuple[tuple[str, Any], ...], int],
    item_id_lookup: dict[str, int],
) -> int | None:
    source_ref = item.get("sourceRef")
    if isinstance(source_ref, dict):
        lookup_key = _source_ref_lookup_key(source_ref)
        if lookup_key is not None and lookup_key in source_ref_lookup:
            return source_ref_lookup[lookup_key]

    item_id = _normalize_item_id(item.get("id"))
    if item_id is not None:
        return item_id_lookup.get(item_id)
    return None


def _find_target_row_for_change(
    target: dict[str, Any],
    source_ref_lookup: dict[tuple[tuple[str, Any], ...], int],
    item_id_lookup: dict[str, int],
) -> int | None:
    source_ref = target.get("sourceRef")
    if isinstance(source_ref, dict):
        lookup_key = _source_ref_lookup_key(source_ref)
        if lookup_key is not None and lookup_key in source_ref_lookup:
            return source_ref_lookup[lookup_key]

    item_id = _normalize_item_id(target.get("itemId"))
    if item_id is not None:
        return item_id_lookup.get(item_id)
    return None


def _find_baseline_item(
    target: dict[str, Any],
    baseline_source_lookup: dict[tuple[tuple[str, Any], ...], dict[str, Any]],
    baseline_item_id_lookup: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    source_ref = target.get("sourceRef")
    if isinstance(source_ref, dict):
        lookup_key = _source_ref_lookup_key(source_ref)
        if lookup_key is not None and lookup_key in baseline_source_lookup:
            return baseline_source_lookup[lookup_key]

    item_id = _normalize_item_id(target.get("itemId"))
    if item_id is not None:
        return baseline_item_id_lookup.get(item_id)
    return None


def _register_row_lookup(
    source_ref_lookup: dict[tuple[tuple[str, Any], ...], int],
    item_id_lookup: dict[str, int],
    item: dict[str, Any],
    row_index: int,
) -> None:
    source_ref = item.get("sourceRef")
    if isinstance(source_ref, dict):
        lookup_key = _source_ref_lookup_key(source_ref)
        if lookup_key is not None:
            source_ref_lookup[lookup_key] = row_index

    item_id = _normalize_item_id(item.get("id"))
    if item_id is not None:
        item_id_lookup[item_id] = row_index


def _source_ref_lookup_key(source_ref: dict[str, Any] | None) -> tuple[tuple[str, Any], ...] | None:
    if not isinstance(source_ref, dict):
        return None
    pairs = []
    for key in SOURCE_REF_COLUMNS:
        value = _normalize_lookup_value(source_ref.get(key))
        if value is not None:
            pairs.append((key, value))
    return tuple(pairs) or None


def _normalize_lookup_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        trimmed = value.strip()
        return trimmed or None
    return value


def _normalize_item_id(value: Any) -> str | None:
    normalized = _normalize_lookup_value(value)
    if normalized is None:
        return None
    return str(normalized)


def _item_to_excel_values(item: dict[str, Any], warnings: list[str], *, include_source_ref: bool) -> dict[str, Any]:
    row_values: dict[str, Any] = {}
    for key, value in item.items():
        if key == "sourceRef":
            if include_source_ref and isinstance(value, dict):
                for source_key in SOURCE_REF_COLUMNS:
                    if source_key in value:
                        row_values[source_key] = _serialize_cell_value(value[source_key], key=source_key, warnings=warnings)
            continue
        if key == "comments":
            if "comment" in item:
                continue
            row_values["comment"] = _serialize_cell_value(value, key="comment", warnings=warnings)
            continue
        row_values[key] = _serialize_cell_value(value, key=key, warnings=warnings)
    return row_values


def _serialize_cell_value(value: Any, *, key: str, warnings: list[str]) -> Any:
    if value is None:
        return None
    if _looks_like_legacy_comments_array(value):
        return _serialize_legacy_comments_array(value, key=key, warnings=warnings)
    if isinstance(value, list):
        return "; ".join(str(part).strip() for part in value if str(part).strip())
    if _looks_like_comment_threads(value):
        return _serialize_comment_threads(value, key=key, warnings=warnings)
    if isinstance(value, (datetime, date)):
        if isinstance(value, date) and not isinstance(value, datetime):
            value = datetime(value.year, value.month, value.day, tzinfo=UTC)
        return _isoformat_datetime(value)
    return value


def _looks_like_comment_threads(value: Any) -> bool:
    return isinstance(value, dict) and isinstance(value.get("threads"), list)


def _looks_like_legacy_comments_array(value: Any) -> bool:
    return isinstance(value, list) and any(isinstance(entry, dict) and ("message" in entry or "text" in entry) for entry in value)


def _serialize_comment_threads(value: dict[str, Any], *, key: str, warnings: list[str]) -> str:
    entries: list[str] = []
    threads = value.get("threads", [])
    for thread in threads:
        if not isinstance(thread, dict):
            warnings.append(f"Skipping malformed thread in field '{key}'")
            continue
        for message in thread.get("messages", []):
            if not isinstance(message, dict):
                warnings.append(f"Skipping malformed message in field '{key}'")
                continue
            timestamp = str(message.get("timestamp", "")).strip()
            rendered_timestamp = _render_comment_timestamp(timestamp)
            author = str(message.get("author", "")).strip()
            text = str(message.get("text", "")).strip()
            entries.append(f"[{rendered_timestamp}] <{author}>: {text}")
    return "; ".join(entries)


def _serialize_legacy_comments_array(value: list[Any], *, key: str, warnings: list[str]) -> str:
    threads: list[dict[str, Any]] = []
    for index, entry in enumerate(value, start=1):
        if not isinstance(entry, dict):
            warnings.append(f"Skipping malformed legacy comment entry in field '{key}'")
            continue

        messages: list[dict[str, Any]] = []
        primary_text = str(entry.get("message") or entry.get("text") or "").strip()
        if primary_text:
            messages.append(
                {
                    "author": str(entry.get("author") or "").strip(),
                    "timestamp": str(entry.get("timestamp") or "").strip(),
                    "text": primary_text,
                }
            )

        replies = entry.get("replies") if isinstance(entry.get("replies"), list) else []
        for reply in replies:
            if not isinstance(reply, dict):
                warnings.append(f"Skipping malformed legacy comment reply in field '{key}'")
                continue
            reply_text = str(reply.get("message") or reply.get("text") or "").strip()
            if not reply_text:
                continue
            messages.append(
                {
                    "author": str(reply.get("author") or "").strip(),
                    "timestamp": str(reply.get("timestamp") or "").strip(),
                    "text": reply_text,
                }
            )

        if messages:
            threads.append({"id": str(entry.get("id") or f"t{index}"), "messages": messages})

    return _serialize_comment_threads({"threads": threads}, key=key, warnings=warnings)


def _render_comment_timestamp(timestamp: str) -> str:
    cleaned_timestamp = str(timestamp or "").strip()
    if not cleaned_timestamp:
        return ""

    normalized_timestamp = _normalize_timestamp_string(cleaned_timestamp)
    if normalized_timestamp:
        return normalized_timestamp[:10]

    return cleaned_timestamp[:10] if len(cleaned_timestamp) >= 10 else cleaned_timestamp


def _ordered_field_names(baseline: dict[str, Any], proposed: dict[str, Any]) -> list[str]:
    field_names = list(proposed.keys())
    for key in baseline.keys():
        if key not in proposed:
            field_names.append(key)
    return field_names


def _copy_target(target: dict[str, Any]) -> dict[str, Any]:
    review_target: dict[str, Any] = {}
    item_id = _normalize_item_id(target.get("itemId"))
    if item_id is not None:
        review_target["itemId"] = item_id
    source_ref = target.get("sourceRef")
    if isinstance(source_ref, dict):
        copied_source_ref = {
            key: _normalize_lookup_value(source_ref.get(key)) for key in SOURCE_REF_COLUMNS if _normalize_lookup_value(source_ref.get(key)) is not None
        }
        if copied_source_ref:
            review_target["sourceRef"] = copied_source_ref
    return review_target


def _read_sheet_value(
    sheet: Worksheet,
    header_index: dict[str, int],
    row_index: int | None,
    field_name: str,
    *,
    warnings: list[str],
) -> Any:
    if row_index is None:
        return None
    column = header_index.get(field_name)
    if column is None:
        return None
    return _normalize_excel_value(
        sheet.cell(row=row_index, column=column).value,
        field_name=field_name,
        row_index=row_index,
        warnings=warnings,
    )


def _next_data_row(sheet: Worksheet) -> int:
    target_row = sheet.max_row + 1 if sheet.max_row >= 1 else 2
    return 2 if target_row == 1 else target_row